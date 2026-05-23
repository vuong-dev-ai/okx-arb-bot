import os, time, threading
from datetime import datetime

import openpyxl
from flask import Flask, jsonify, render_template

from strategy import (
    get_funding_rates, get_available_usdt,
    open_position, close_position,
    check_exit_conditions, estimate_pnl,
    MIN_FUNDING_RATE, POSITION_PCT, MIN_USDT,
)
from excel_logger import log_pnl_snapshot, export_json, push_to_github, EXCEL_FILE

app = Flask(__name__)

_lock  = threading.Lock()
_state = {
    'running':       False,
    'positions':     [],
    'opportunities': [],
    'usdt':          0.0,
    'log':           [],
    'last_update':   '-',
}

MAX_POS   = 3
SCAN_INT  = 300      # scan funding rate mỗi 5 phút
MON_INT   = 30       # cập nhật PnL mỗi 30 giây
EXCL_INT  = 8*3600   # ghi Excel mỗi 8 giờ
PUSH_INT  = 300      # push GitHub Pages mỗi 5 phút
TICK      = 5        # vòng lặp chính mỗi 5 giây (để dừng nhanh)
MAX_LOG   = 300
_bot_thread = None


def _log(msg: str):
    line = f"[{datetime.now().strftime('%H:%M:%S')}] {msg}"
    with _lock:
        _state['log'].append(line)
        if len(_state['log']) > MAX_LOG:
            _state['log'] = _state['log'][-MAX_LOG:]


def _bot():
    last_scan = last_mon = last_excel = last_push = 0
    last_opps: list = []

    _log("━━━ Bot OKX Funding Arb khởi động ━━━")
    with _lock:
        _state['usdt'] = get_available_usdt()
    _log(f"Số dư: ${_state['usdt']:.2f} USDT")

    while True:
        with _lock:
            if not _state['running']:
                break
        now = time.time()

        # ── Cập nhật PnL & kiểm tra thoát ───────────────────────
        if now - last_mon >= MON_INT:
            with _lock:
                positions = list(_state['positions'])

            for p in positions:
                try:
                    p['_pnl']  = estimate_pnl(p)
                    ok, rate   = check_exit_conditions(p)
                    p['_exit'] = ok
                except Exception as e:
                    _log(f"[{p['coin']}] Lỗi cập nhật: {e}")

            for p in [x for x in positions if x.get('_exit')]:
                _log(f"[{p['coin']}] Funding âm → đóng vị thế...")
                try:
                    close_position(p)
                    with _lock:
                        _state['positions'] = [x for x in _state['positions']
                                               if x['coin'] != p['coin']]
                    _log(f"[{p['coin']}] Đóng ✓")
                except Exception as e:
                    _log(f"[{p['coin']}] Lỗi đóng: {e}")

            with _lock:
                _state['usdt']        = get_available_usdt()
                _state['last_update'] = datetime.now().strftime('%H:%M:%S')
            last_mon = now

        # ── Scan cơ hội ───────────────────────────────────────────
        if now - last_scan >= SCAN_INT:
            with _lock:
                n_pos = len(_state['positions'])

            if n_pos < MAX_POS:
                _log(f"SCAN — ${_state['usdt']:.2f} USDT  |  {datetime.now().strftime('%H:%M')}")
                opps = get_funding_rates()
                last_opps = opps or []
                with _lock:
                    _state['opportunities'] = last_opps[:10]

                if opps:
                    with _lock:
                        open_coins = {p['coin'] for p in _state['positions']}
                    for opp in opps:
                        with _lock:
                            n_pos = len(_state['positions'])
                            go    = _state['running']
                        if not go or n_pos >= MAX_POS:
                            break
                        if opp['coin'] in open_coins:
                            continue
                        if opp['funding_rate'] < MIN_FUNDING_RATE:
                            break
                        usdt   = get_available_usdt()
                        amount = usdt * POSITION_PCT
                        if amount < MIN_USDT:
                            _log(f"Số dư thấp (${usdt:.2f}) — dừng scan")
                            break
                        _log(f"[{opp['coin']}] Vào lệnh ${amount:.2f} @ {opp['funding_rate']*100:.4f}%/8h")
                        pos = open_position(opp, amount)
                        if pos:
                            with _lock:
                                _state['positions'].append(pos)
                            open_coins.add(opp['coin'])
                            _log(f"[{opp['coin']}] Mở ✓ giá=${pos['entry_price']:.2f}")
                else:
                    _log("Không lấy được funding rate")
            last_scan = now

        # ── Ghi Excel mỗi 8 giờ ──────────────────────────────────
        if now - last_excel >= EXCL_INT:
            with _lock:
                ps = list(_state['positions'])
                u  = _state['usdt']
            if ps:
                pl = [estimate_pnl(p) for p in ps]
                try:
                    log_pnl_snapshot(ps, pl, u)
                    _log("Ghi Excel ✓ → pnl_log.xlsx")
                except Exception as e:
                    _log(f"Lỗi ghi Excel: {e}")
            last_excel = now

        # ── Push GitHub Pages mỗi 5 phút ─────────────────────────
        if now - last_push >= PUSH_INT:
            with _lock:
                ps = list(_state['positions'])
                u  = _state['usdt']
            try:
                pl = [p.get('_pnl') or estimate_pnl(p) for p in ps]
                export_json({'positions': ps, 'pnl_list': pl,
                             'opportunities': last_opps, 'usdt': u})
                if push_to_github():
                    _log("GitHub Pages ✓ (data.json)")
            except Exception as e:
                _log(f"Lỗi push GitHub: {e}")
            last_push = now

        time.sleep(TICK)

    # ── Đóng tất cả khi dừng ─────────────────────────────────────
    with _lock:
        ps = list(_state['positions'])
    if ps:
        _log(f"Đóng {len(ps)} vị thế...")
        for p in ps:
            try:
                close_position(p)
                with _lock:
                    _state['positions'] = [x for x in _state['positions']
                                           if x['coin'] != p['coin']]
                _log(f"[{p['coin']}] Đóng ✓")
            except Exception as e:
                _log(f"[{p['coin']}] Lỗi: {e}")
    _log("━━━ Bot đã dừng ━━━")


# ── API ───────────────────────────────────────────────────────────

@app.route('/api/status')
def api_status():
    with _lock:
        ps      = list(_state['positions'])
        opps    = list(_state['opportunities'])
        usdt    = _state['usdt']
        running = _state['running']
        upd     = _state['last_update']
        logs    = list(_state['log'][-120:])

    out_ps = []
    for p in ps:
        pnl = p.get('_pnl') or {}
        vin = p['contracts'] * p['ct_val'] * p['entry_price']
        pct = (pnl.get('total_pnl', 0) / vin * 100) if vin else 0
        out_ps.append({
            'coin':        p['coin'],
            'entry_price': p['entry_price'],
            'cur_price':   pnl.get('price'),
            'usdt_in':     round(vin, 2),
            'funding_pnl': round(pnl.get('funding_pnl', 0), 4),
            'price_pnl':   round(pnl.get('price_pnl', 0), 4),
            'total_pnl':   round(pnl.get('total_pnl', 0), 4),
            'pct':         round(pct, 3),
            'n_pay':       pnl.get('n_payments', 0),
            'open_time':   datetime.fromtimestamp(p['open_time']).strftime('%d/%m %H:%M'),
            'exit':        p.get('_exit', False),
        })

    return jsonify({
        'running': running, 'usdt': usdt, 'positions': out_ps,
        'opps':    [{'coin': o['coin'], 'rate': o['funding_rate'],
                     'apy': o['annualized'], 'next': o['next_rate']} for o in opps],
        'logs': logs, 'last_update': upd,
    })


@app.route('/api/start', methods=['POST'])
def api_start():
    global _bot_thread
    with _lock:
        if _state['running']:
            return jsonify({'ok': False, 'msg': 'Bot đang chạy rồi'})
        _state['running'] = True
    _bot_thread = threading.Thread(target=_bot, daemon=True)
    _bot_thread.start()
    return jsonify({'ok': True})


@app.route('/api/stop', methods=['POST'])
def api_stop():
    with _lock:
        if not _state['running']:
            return jsonify({'ok': False, 'msg': 'Bot chưa chạy'})
        _state['running'] = False
    return jsonify({'ok': True})


@app.route('/api/close/<coin>', methods=['POST'])
def api_close(coin):
    with _lock:
        pos = next((p for p in _state['positions'] if p['coin'] == coin), None)
    if not pos:
        return jsonify({'ok': False, 'msg': f'Không tìm thấy {coin}'})

    def _do():
        _log(f"[{coin}] Đóng thủ công...")
        try:
            close_position(pos)
            with _lock:
                _state['positions'] = [p for p in _state['positions'] if p['coin'] != coin]
            _log(f"[{coin}] Đóng thủ công ✓")
        except Exception as e:
            _log(f"[{coin}] Lỗi: {e}")
    threading.Thread(target=_do, daemon=True).start()
    return jsonify({'ok': True})


@app.route('/api/pnl-history')
def api_pnl():
    excel_path = os.path.join(os.path.dirname(__file__), EXCEL_FILE)
    if not os.path.exists(excel_path):
        return jsonify({'headers': [], 'rows': []})
    try:
        wb   = openpyxl.load_workbook(excel_path, data_only=True)
        ws   = wb.active
        hdrs = [c.value for c in ws[1] if c.value]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None and v != '' for v in row):
                rows.append([str(v) if v is not None else '' for v in row[:len(hdrs)]])
        wb.close()
        return jsonify({'headers': hdrs, 'rows': rows})
    except Exception as e:
        return jsonify({'error': str(e), 'headers': [], 'rows': []})


@app.route('/')
def index():
    return render_template('index.html')


if __name__ == '__main__':
    print("\n" + "="*50)
    print("  OKX Arb Bot — Web Dashboard")
    print("  Mở trình duyệt: http://localhost:5000")
    print("="*50 + "\n")
    app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)
