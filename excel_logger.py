import os
import json
import subprocess
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EXCEL_FILE = "pnl_log.xlsx"
_BASE      = os.path.dirname(os.path.abspath(__file__))
DOCS_DIR   = os.path.join(_BASE, "docs")

HEADERS = [
    "Thời gian", "Coin", "Giá vào ($)", "Giá hiện tại ($)",
    "Vốn vào ($)", "Funding PnL ($)", "Giá PnL ($)",
    "Tổng PnL ($)", "% Lãi", "Số lần nhận funding", "Số dư USDT ($)",
]
COL_WIDTHS = [20, 8, 14, 16, 13, 15, 13, 14, 9, 21, 16]


# ── Excel ─────────────────────────────────────────────────────────

def _get_or_create_wb():
    path = os.path.join(_BASE, EXCEL_FILE)
    if os.path.exists(path):
        return openpyxl.load_workbook(path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lời Lỗ"
    _write_header(ws)
    wb.save(path)
    return wb


def _write_header(ws):
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def log_pnl_snapshot(positions: list, pnl_list: list, usdt_balance: float):
    wb = _get_or_create_wb()
    ws = wb["Lời Lỗ"]
    if ws.cell(1, 1).value != HEADERS[0]:
        _write_header(ws)

    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    valid   = [(p, pnl) for p, pnl in zip(positions, pnl_list) if pnl is not None]
    if not valid:
        return

    for i, (pos, pnl) in enumerate(valid):
        row    = ws.max_row + 1
        usdt_in = pos['contracts'] * pos['ct_val'] * pos['entry_price']
        pct    = (pnl['total_pnl'] / usdt_in * 100) if usdt_in > 0 else 0
        is_last = (i == len(valid) - 1)

        values = [
            now_str, pos['coin'],
            round(pos['entry_price'], 4), round(pnl['price'], 4),
            round(usdt_in, 2),
            round(pnl['funding_pnl'], 4), round(pnl['price_pnl'], 4),
            round(pnl['total_pnl'], 4), round(pct, 3),
            pnl['n_payments'],
            round(usdt_balance, 2) if is_last else "",
        ]
        green    = pnl['total_pnl'] >= 0
        pnl_fill = PatternFill("solid", fgColor="C6EFCE" if green else "FFC7CE")
        pnl_clr  = "276221" if green else "9C0006"
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            if col in (6, 7, 8):
                cell.fill = pnl_fill
                cell.font = Font(color=pnl_clr, bold=(col == 8))

    if len(valid) > 1:
        row        = ws.max_row + 1
        total_pnl  = sum(pnl['total_pnl'] for _, pnl in valid)
        total_usdt = sum(p['contracts'] * p['ct_val'] * p['entry_price'] for p, _ in valid)
        total_pct  = (total_pnl / total_usdt * 100) if total_usdt > 0 else 0
        sfill      = PatternFill("solid", fgColor="D9E1F2")
        green      = total_pnl >= 0
        summary    = [now_str, "TỔNG", "", "", round(total_usdt, 2), "", "",
                      round(total_pnl, 4), round(total_pct, 3), "", round(usdt_balance, 2)]
        for col, val in enumerate(summary, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = sfill
            cell.alignment = Alignment(horizontal="center")
            if col == 2: cell.font = Font(bold=True)
            if col == 8: cell.font = Font(bold=True, color="276221" if green else "9C0006")

    ws.append([""] * len(HEADERS))
    wb.save(os.path.join(_BASE, EXCEL_FILE))


# ── GitHub Pages export ───────────────────────────────────────────

def export_json(live: dict = None):
    """
    live = {
        'positions': list[dict],   # dict từ open_position(), có thể có _pnl
        'pnl_list':  list[dict],   # kết quả estimate_pnl()
        'opportunities': list[dict],
        'usdt': float,
    }
    """
    os.makedirs(DOCS_DIR, exist_ok=True)

    # ── Live data ─────────────────────────────────────────────────
    live_positions = []
    total_live_pnl = 0.0
    top_rates      = []
    usdt_now       = 0.0

    if live:
        usdt_now = live.get('usdt', 0)
        positions = live.get('positions', [])
        pnl_list  = live.get('pnl_list', [])
        for pos, pnl in zip(positions, pnl_list):
            if not pnl:
                continue
            vin = pos['contracts'] * pos['ct_val'] * pos['entry_price']
            pct = (pnl['total_pnl'] / vin * 100) if vin else 0
            total_live_pnl += pnl['total_pnl']
            live_positions.append({
                'coin':          pos['coin'],
                'entry_price':   round(pos['entry_price'], 2),
                'current_price': round(pnl['price'], 2) if pnl.get('price') else None,
                'usdt_in':       round(vin, 2),
                'funding_pnl':   round(pnl['funding_pnl'], 4),
                'price_pnl':     round(pnl['price_pnl'], 4),
                'total_pnl':     round(pnl['total_pnl'], 4),
                'pct':           round(pct, 3),
                'n_payments':    pnl['n_payments'],
                'open_time':     datetime.fromtimestamp(pos['open_time']).strftime('%d/%m %H:%M'),
            })
        for o in live.get('opportunities', [])[:8]:
            top_rates.append({
                'coin': o['coin'],
                'rate': o['funding_rate'],
                'apy':  o['annualized'],
                'next': o['next_rate'],
            })

    # ── Excel history ─────────────────────────────────────────────
    excel_path = os.path.join(_BASE, EXCEL_FILE)
    hdrs, rows = [], []
    chart      = {'labels': [], 'values': []}
    last_bal   = ""

    if os.path.exists(excel_path):
        wb   = openpyxl.load_workbook(excel_path, data_only=True)
        ws   = wb.active
        hdrs = [c.value for c in ws[1] if c.value]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None and v != "" for v in row):
                rows.append([str(v) if v is not None else "" for v in row[:len(hdrs)]])
        wb.close()

        coin_i = hdrs.index("Coin")            if "Coin"            in hdrs else -1
        pnl_i  = hdrs.index("Tổng PnL ($)")   if "Tổng PnL ($)"   in hdrs else -1
        time_i = hdrs.index("Thời gian")       if "Thời gian"       in hdrs else -1
        bal_i  = hdrs.index("Số dư USDT ($)")  if "Số dư USDT ($)" in hdrs else -1

        sum_rows = [r for r in rows if coin_i >= 0 and r[coin_i] == "TỔNG"]
        if not sum_rows:
            sum_rows = [r for r in rows if coin_i >= 0 and r[coin_i] not in ("", "TỔNG")]

        cum = 0.0
        for r in sum_rows:
            v = float(r[pnl_i]) if pnl_i >= 0 and r[pnl_i] else 0
            cum += v
            chart['labels'].append(r[time_i] if time_i >= 0 else "")
            chart['values'].append(round(cum, 4))

        if bal_i >= 0:
            for r in reversed(rows):
                if r[bal_i]:
                    last_bal = r[bal_i]
                    break

    data = {
        "updated_at":    datetime.now().strftime("%d/%m/%Y %H:%M"),
        "usdt_balance":  usdt_now,
        "total_live_pnl": round(total_live_pnl, 4),
        "positions":     live_positions,
        "top_rates":     top_rates,
        "history": {
            "headers":       hdrs,
            "rows":          rows[-200:],   # giới hạn 200 dòng
            "chart":         chart,
            "last_balance":  last_bal,
            "total_periods": len(chart['labels']),
            "total_pnl":     chart['values'][-1] if chart['values'] else 0,
        },
    }

    with open(os.path.join(DOCS_DIR, "data.json"), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def push_to_github():
    """Git commit + push docs/data.json."""
    try:
        subprocess.run(["git", "-C", _BASE, "add", "docs/data.json"],
                       check=True, capture_output=True)
        res = subprocess.run(
            ["git", "-C", _BASE, "commit", "-m",
             f"data: {datetime.now().strftime('%d/%m %H:%M')}"],
            capture_output=True, text=True,
        )
        if "nothing to commit" in (res.stdout + res.stderr):
            return True
        subprocess.run(["git", "-C", _BASE, "push"],
                       check=True, capture_output=True)
        return True
    except subprocess.CalledProcessError:
        return False
